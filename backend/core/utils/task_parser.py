from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import json


filename = "Тестовый файл.xlsx"
wb = load_workbook(filename=filename)
sheets = [sheet for sheet in wb]
result = {}
for sheet in sheets:
    i =1
    while sheet.cell(row=i, column=1).value:
        if not sheet.title in result.keys():
            result[sheet.title] = {"Задание": "", "Ответ": ""}
        result[sheet.title]['Задание'] = sheet.cell(row=i, column=1).value
        result[sheet.title]['Ответ'] = sheet.cell(row=i, column=3).value
        i += 1


json_res = json.dumps(result, ensure_ascii=False, indent=4)

print(json_res)

with open("result.json", "w", encoding="utf-8") as f:
  f.write(json_res)